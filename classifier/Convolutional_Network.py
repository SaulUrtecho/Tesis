# **PARTE 1: EN ESTA PARTE SE CONSTRUYE LA RED NEURONAL CONVOLUCIONAL**
from keras.models import Sequential
from keras.layers import Flatten, Dense, Dropout
from keras.layers.convolutional import Conv2D, MaxPooling2D
from keras.preprocessing.image import ImageDataGenerator
from keras import backend as K
import matplotlib.pyplot as plt
from openpyxl import Workbook


TRAIN_PATH = 'C:/Users/saulu/Documents/binary_classifier/classifier/conjunto_de_datos/conjunto_de_entrenamiento'
TEST_PATH = 'C:/Users/saulu/Documents/binary_classifier/classifier/conjunto_de_datos/conjunto_de_pruebas'

K.clear_session() # limpia el backend de keras

# Debido a la poca cantidad de imagenes de muestra, se utilzan GENERADORES 
# los cuales tienen la funcion de a partir de una imagen usar transformaciones
# geometricas para aumentar el conjunto de entrenamiento
# las transformaciones solo aplican para el conjunto de entrenamiento,
# en el de validacion solo se convierte a valor flotante

train_datagen = ImageDataGenerator(
        rescale=1./255,
        shear_range=0.3,
        zoom_range=0.3,
        horizontal_flip=True)

test_datagen = ImageDataGenerator(rescale=1./255)

training_set = train_datagen.flow_from_directory(
        TRAIN_PATH,
        target_size=(200, 200),
        batch_size=32,
        class_mode='binary')

test_set = test_datagen.flow_from_directory(
        TEST_PATH,
        target_size=(200, 200),
        batch_size=32,
        class_mode='binary')

print(test_set.class_indices) # etiquetas tienen las clases [0-Enfermo] o [1-Sano]

print(training_set.class_indices)

# ==== Calculando los pasos de Entrenamiento y validacion ====
# se calcula diviendo el tamaño de cada conjunto entre el tamaño de lote
pasos_train = training_set.n//training_set.batch_size
pasos_val = test_set.n//test_set.batch_size
# Creando la Red Neuronal Convolucional
model = Sequential()
# Step - Convolution
# Si la imagen es mayor a 128px usar un kernel size mayor a 3x3
model.add(Conv2D(
        filters=32,
        kernel_size=5,
        strides=(1, 1),
        padding='same',
        activation='relu',
        input_shape=(200, 200, 3)))

model.add(MaxPooling2D(
        pool_size=(2, 2)))
# Adding a second convolutional layer
model.add(Conv2D(
        filters=64,
        kernel_size=3,
        strides=(1, 1),
        padding='same',
        activation='relu'))

model.add(MaxPooling2D(
        pool_size=(2, 2)))
# Step - Flattening
model.add(Flatten()) # los convertimos a un vector lineal
# Step - Full connection, red tipo perceptron con 128 neuronas
model.add(Dense(units=128, activation='relu'))

""" EL Overfitting es cuando el modelo solo toma como validos los datos
con los cuales se entreno y no logra generalizar con datos que sean
diferentes a la base de datos inicial """

# El dropout establece aleatoriamente entradas en 0 
# en cada paso de entrenamiento lo que ayuda a evitar el sobreajuste
model.add(Dropout(0.5))

# Se establece una sola neurona para la capa de salida en vez de dos
# La salida 0 (<0.5) es ENFERMO y 1 (>=0.5) es SANO
# Se utiliza la funcion de activacion sigmoidal, esta funcion existe entre 0 y 1
# la funcion sigmoidal tiene una derivada sencilla
# se usa en modelos cuya probabilidad este entre 0 y 1
model.add(Dense(units=1, activation='sigmoid'))

# Compiling the CNN
# La funcion de perdida nos informa que tan precisa es nuestra red
# La pérdida de entropía cruzada se utiliza al ajustar los pesos del modelo
# durante el entrenamiento. El objetivo es minimizar la pérdida, es decir,
# cuanto menor sea la pérdida, mejor será el modelo. Calcula la pérdida de
#  entropía cruzada entre etiquetas verdaderas y etiquetas predichas.
model.compile(optimizer='adam', loss='binary_crossentropy', metrics=['accuracy'])

# Part 2 - Fitting the CNN to the images
# se le asignan los parametros al modelo para que pueda entrenarse
# y se almacenan en la variable history
history = model.fit(
        training_set,
        steps_per_epoch=pasos_train,
        epochs=15,
        validation_data=test_set,
        validation_steps=pasos_val)

# --- Creacion del directorio donde se guardara el modelo ---
# directorio ='./DATOS_RED/'
# si el directorio no existe se crea
# if not os.path.exists(directorio):
# os.mkdir(directorio)  # se crea la carpeta en la ruta actual del proyecto
# el modelo y los pesos son guardados respectivamente
model.save('./DATOS_RED/Modelo.h5')

model.save_weights('./DATOS_RED/Pesos.h5')

# Con el objeto creado podemos acceder al diccionario donde se
# almacenan los parametros que arroja el modelo
history_dict = history.history
print(history_dict.keys())

# Estos son las claves del diccionario obtenido de history_dict.keys(),
# con ellos obtenemos los valores de cada clave que se generaron
# por cada epoca de entrenamiento
acc = history.history['accuracy']
val_acc = history.history['val_accuracy'] # presicion FINAL del modelo!
loss = history.history['loss']
val_loss = history.history['val_loss']  

# Procedemos a mostrar los datos obtenidos durante el entrenamiento
print(acc)  
print(val_acc) 
print(loss)
print(val_loss) 
print("----------------- RESULTADO DEL MODELO -----------------")
print()
print()
print()
print()
# mostramos el valor de la presicion almacenado en la ultima posicion
# de la lista del valor en la key val_accuracy
print("La Presición FINAL del modelo es: ", history_dict['val_accuracy'][-1])

print("----------------- GRÁFICAS -----------------")
print()
print()
print()
print()

epochs = range(1, len(loss) + 1)
plt.plot(epochs, loss, 'y', label='Training loss')
plt.plot(epochs, val_loss, 'r', label='Validation loss')
plt.title('Training and validation loss')
plt.xlabel('Epochs')
plt.ylabel('Loss')
plt.legend()
plt.savefig('./DATOS_RED/Perdidas.png')
plt.show()

plt.plot(epochs, acc, 'y', label='Training acc')
plt.plot(epochs, val_acc, 'g', label='Validation acc')
plt.title('Training and validation accuracy')
plt.xlabel('Epochs')
plt.ylabel('Accuracy')
plt.legend()
plt.savefig('./DATOS_RED/Precisiones.png')
plt.show()

# Aqui se crea el archivo excel donde se almacena los datos del entrenamiento
wb = Workbook()

hoja = wb.active
hoja.title = "registro del Entrenamiento"

num = 0
fila = 2
accuary = 2
val_ac = 3
val_l = 4
lss = 5

for accu, vala, vals, loss in zip(acc, val_acc, val_loss, loss):
    hoja.cell(row=1, column=2, value='Accuracy')
    hoja.cell(row=1, column=3, value='Val_accuracy')
    hoja.cell(row=1, column=4, value='Val_loss')
    hoja.cell(row=1, column=5, value='Loss')

    hoja.cell(column=1, row=fila, value=num)
    hoja.cell(column=accuary, row=fila, value= accu)
    hoja.cell(column=val_ac, row=fila, value= vala)
    hoja.cell(column=val_l, row=fila, value= vals)
    hoja.cell(column=lss, row=fila, value= loss)
    fila += 1
    num += 1

wb.save('./DATOS_RED/Resultados.xlsx') # los datos son guardados